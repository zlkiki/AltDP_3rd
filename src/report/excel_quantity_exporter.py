"""Multi-sheet Excel Quantity Exporter using openpyxl (Phase 17-2).

Generates 3-sheet professional Excel workbooks:
Sheet 1: Summary (총괄 물량 집계표)
Sheet 2: By Story (층별 물량 집계표)
Sheet 3: Member Breakdown (부재별 상세 내역서)
"""

from typing import Optional, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.engine.project.quantity_engine import ProjectQuantitySummary


class ExcelQuantityExporter:
    """Exports structural quantities into a styled multi-sheet Excel file."""

    # Styling Palette
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    WHITE_FONT = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    BOLD_FONT = Font(name="Malgun Gothic", size=10, bold=True, color="000000")
    REGULAR_FONT = Font(name="Malgun Gothic", size=10, color="000000")
    TITLE_FONT = Font(name="Malgun Gothic", size=14, bold=True, color="1F4E78")

    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    @classmethod
    def export_to_bytes(cls, summary: ProjectQuantitySummary) -> BytesIO:
        """Create multi-sheet workbook and return as BytesIO stream."""
        wb = openpyxl.Workbook()
        
        # 1. Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "총괄 물량 집계표"
        cls._write_summary_sheet(ws1, summary)

        # 2. Sheet 2: By Story
        ws2 = wb.create_sheet(title="층별 물량 집계표")
        cls._write_story_sheet(ws2, summary)

        # 3. Sheet 3: Member Breakdown
        ws3 = wb.create_sheet(title="부재별 상세 내역서")
        cls._write_member_sheet(ws3, summary)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @classmethod
    def _write_summary_sheet(cls, ws: Any, s: ProjectQuantitySummary):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value="KDS 표준 총괄 공사 물량 집계표 (Summary)").font = cls.TITLE_FONT

        headers = ["공종 / 구분", "규격 및 항목", "단위", "수량 / 중량", "비고"]
        for col_idx, h in enumerate(headers, start=2):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.fill = cls.HEADER_FILL
            c.font = cls.WHITE_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")

        rows_data = [
            ("레미콘 (Concrete)", "fck = 24 MPa 표준", "m³", s.total_concrete_vol_m3, "타설 체적"),
            ("거푸집 (Formwork)", "합판/유로폼 기준", "m²", s.total_formwork_area_m2, "접촉 표면적"),
            ("철근 총 중량 (Rebar)", "HD (SD400/500)", "ton", s.total_rebar_weight_ton, "정착/이음 포함"),
        ]

        # Rebar breakdown by size
        for sz, wt in sorted(s.rebar_totals_by_size_kg.items()):
            rows_data.append(("철근 소계", f"규격 {sz}", "ton", round(wt / 1000.0, 3), f"{wt:.1f} kg"))

        curr_r = 5
        for item in rows_data:
            for col_idx, val in enumerate(item, start=2):
                c = ws.cell(row=curr_r, column=col_idx, value=val)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="center")
            curr_r += 1

        cls._auto_fit_columns(ws)

    @classmethod
    def _write_story_sheet(cls, ws: Any, s: ProjectQuantitySummary):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value="층별 공사 물량 집계표 (By Story)").font = cls.TITLE_FONT

        headers = ["층 (Story)", "콘크리트 체적 (m³)", "거푸집 면적 (m²)", "철근 총 중량 (ton)"]
        for col_idx, h in enumerate(headers, start=2):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.fill = cls.HEADER_FILL
            c.font = cls.WHITE_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")

        curr_r = 5
        for story, vals in sorted(s.story_breakdowns.items()):
            row_vals = [story, vals["concrete_m3"], vals["formwork_m2"], vals["rebar_ton"]]
            for col_idx, val in enumerate(row_vals, start=2):
                c = ws.cell(row=curr_r, column=col_idx, value=val)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="center")
            curr_r += 1

        cls._auto_fit_columns(ws)

    @classmethod
    def _write_member_sheet(cls, ws: Any, s: ProjectQuantitySummary):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value="부재별 상세 물량 내역서 (Member Breakdown)").font = cls.TITLE_FONT

        headers = ["부재 ID", "층 (Story)", "부재 구분", "콘크리트 (m³)", "거푸집 (m²)", "철근 (kg)"]
        for col_idx, h in enumerate(headers, start=2):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.fill = cls.HEADER_FILL
            c.font = cls.WHITE_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")

        curr_r = 5
        for m in s.member_details:
            row_vals = [m.member_id, m.story, m.member_type, m.concrete_vol_m3, m.formwork_area_m2, m.rebar_weight_kg]
            for col_idx, val in enumerate(row_vals, start=2):
                c = ws.cell(row=curr_r, column=col_idx, value=val)
                c.font = cls.REGULAR_FONT
                c.border = cls.THIN_BORDER
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="center")
            curr_r += 1

        cls._auto_fit_columns(ws)

    @classmethod
    def _auto_fit_columns(cls, ws: Any):
        """Auto-adjust column widths with padding."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
