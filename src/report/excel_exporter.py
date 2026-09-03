"""Excel Structural Calculation Report Exporter for AltDP_3rd.

Adheres to Midas Design+ CMSExcel format. Utilizes openpyxl to generate
professional multi-sheet structural design calculation workbooks.
"""

import io
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportExporter:
    """Multi-sheet Excel Workbook Exporter for Structural Calculation Reports."""

    def __init__(self):
        # Color palettes
        self.header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        self.sub_header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        self.pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        self.warn_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

        self.font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.font_title = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        self.font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
        self.font_regular = Font(name="Calibri", size=10, color="334155")
        self.font_pass = Font(name="Calibri", size=10, bold=True, color="166534")
        self.font_fail = Font(name="Calibri", size=10, bold=True, color="991B1B")

        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_left = Alignment(horizontal="left", vertical="center")
        self.align_right = Alignment(horizontal="right", vertical="center")

        thin_side = Side(border_style="thin", color="CBD5E1")
        self.thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def export_workbook_bytes(
        self,
        project_info: Dict[str, Any],
        member_info: Dict[str, Any],
        material_info: Dict[str, Any],
        section_info: Dict[str, Any],
        loads_info: Dict[str, Any],
        checks: List[Dict[str, Any]],
        summary_dcr: float,
        is_safe: bool,
    ) -> bytes:
        """Create and export full Excel workbook as bytes buffer."""
        wb = openpyxl.Workbook()
        
        # 1. Summary Sheet
        ws_sum = wb.active
        ws_sum.title = "Summary & Overview"
        self._build_summary_sheet(ws_sum, project_info, member_info, summary_dcr, is_safe)

        # 2. Material & Section Sheet
        ws_mat = wb.create_sheet(title="Material & Section")
        self._build_material_section_sheet(ws_mat, material_info, section_info)

        # 3. Detailed Checks Sheet
        ws_chk = wb.create_sheet(title="Detailed Checks")
        self._build_checks_sheet(ws_chk, checks)

        # 4. Loads & Combinations Sheet
        ws_load = wb.create_sheet(title="Design Loads")
        self._build_loads_sheet(ws_load, loads_info)

        # Auto-fit column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _build_summary_sheet(self, ws, project: Dict[str, Any], member: Dict[str, Any], dcr: float, is_safe: bool):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value=f"{project.get('title', 'AltDP_3rd 구조계산서')} - 부재 요약").font = self.font_title
        
        # Meta info
        meta_rows = [
            ("부재 ID (Member ID)", member.get("id", "M-101")),
            ("부재 유형 (Member Type)", member.get("type", "RC Beam")),
            ("적용 설계기준 (Design Code)", project.get("code", "KDS 14 20 00 / KDS 14 31 00")),
            ("설계자 / 검토자", f"{project.get('author', 'AltDP_3rd')} / {project.get('checker', 'Senior PE')}"),
            ("해석 및 생성 일자", project.get("date", "2026-09-02")),
            ("최대 응력비 (Max DCR)", f"{dcr:.3f}"),
            ("종합 설계 판정", "PASS (적합)" if is_safe else "FAIL (부적합)"),
        ]

        for i, (label, val) in enumerate(meta_rows, start=4):
            c_lbl = ws.cell(row=i, column=2, value=label)
            c_lbl.font = self.font_bold
            c_lbl.fill = self.sub_header_fill
            c_lbl.border = self.thin_border
            
            c_val = ws.cell(row=i, column=3, value=val)
            c_val.font = self.font_pass if ("PASS" in str(val) or label.startswith("최대")) else (self.font_fail if "FAIL" in str(val) else self.font_regular)
            if "PASS" in str(val):
                c_val.fill = self.pass_fill
            elif "FAIL" in str(val):
                c_val.fill = self.fail_fill
            c_val.border = self.thin_border

    def _build_material_section_sheet(self, ws, mat: Dict[str, Any], sect: Dict[str, Any]):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value="재료 물성치 및 단면 제원 (Materials & Geometry)").font = self.font_title

        # Materials
        ws.cell(row=4, column=2, value="재료 물성 항목").font = self.font_header
        ws.cell(row=4, column=2).fill = self.header_fill
        ws.cell(row=4, column=3, value="설계치 / 강도").font = self.font_header
        ws.cell(row=4, column=3).fill = self.header_fill

        r = 5
        for k, v in mat.items():
            ws.cell(row=r, column=2, value=str(k)).border = self.thin_border
            ws.cell(row=r, column=3, value=str(v)).border = self.thin_border
            r += 1

        # Section
        r += 1
        ws.cell(row=r, column=2, value="단면 기하 파라미터").font = self.font_header
        ws.cell(row=r, column=2).fill = self.header_fill
        ws.cell(row=r, column=3, value="치수 및 성질").font = self.font_header
        ws.cell(row=r, column=3).fill = self.header_fill
        r += 1
        for k, v in sect.items():
            ws.cell(row=r, column=2, value=str(k)).border = self.thin_border
            ws.cell(row=r, column=3, value=str(v)).border = self.thin_border
            r += 1

    def _build_checks_sheet(self, ws, checks: List[Dict[str, Any]]):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value="상세 한계상태 설계 검토표 (Detailed Limit State Checks)").font = self.font_title

        headers = ["No", "검토 항목 (Check Title)", "계수 소요치 (Demand)", "설계 강도 (Capacity)", "응력비 (DCR)", "판정 (Status)"]
        for c_idx, h in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border

        for r_idx, chk in enumerate(checks, start=5):
            dcr_val = float(chk.get("dcr", 0.0))
            is_ok = dcr_val <= 1.0

            ws.cell(row=r_idx, column=2, value=r_idx - 4).border = self.thin_border
            ws.cell(row=r_idx, column=3, value=chk.get("title", "-")).border = self.thin_border
            ws.cell(row=r_idx, column=4, value=str(chk.get("demand", "-"))).border = self.thin_border
            ws.cell(row=r_idx, column=5, value=str(chk.get("capacity", "-"))).border = self.thin_border
            
            c_dcr = ws.cell(row=r_idx, column=6, value=f"{dcr_val:.3f}")
            c_dcr.border = self.thin_border
            c_dcr.alignment = self.align_right
            c_dcr.font = self.font_pass if is_ok else self.font_fail

            c_status = ws.cell(row=r_idx, column=7, value="OK" if is_ok else "NG")
            c_status.border = self.thin_border
            c_status.alignment = self.align_center
            c_status.fill = self.pass_fill if is_ok else self.fail_fill
            c_status.font = self.font_pass if is_ok else self.font_fail

    def _build_loads_sheet(self, ws, loads: Dict[str, Any]):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=2, column=2, value="계수 설계하중 및 지배 하중조합 (Design Loads)").font = self.font_title

        ws.cell(row=4, column=2, value="하중 성분 / 항목").font = self.font_header
        ws.cell(row=4, column=2).fill = self.header_fill
        ws.cell(row=4, column=3, value="설계 하중값").font = self.font_header
        ws.cell(row=4, column=3).fill = self.header_fill

        r = 5
        for k, v in loads.items():
            ws.cell(row=r, column=2, value=str(k)).border = self.thin_border
            ws.cell(row=r, column=3, value=str(v)).border = self.thin_border
            r += 1

    def export_project_workbook_bytes(
        self,
        project_info: Dict[str, Any],
        members_data: List[Dict[str, Any]],
    ) -> bytes:
        """Create integrated multi-sheet Excel workbook for all project members."""
        wb = openpyxl.Workbook()

        # Sheet 1: Project_Summary
        ws_sum = wb.active
        ws_sum.title = "Project_Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        ws_sum.cell(row=2, column=2, value=f"{project_info.get('title', 'AltDP_3rd')} - 전 부재 설계 요약표").font = self.font_title

        headers = ["No", "부재 ID", "부재 종류", "단면 제원", "최대 DCR", "안전성 판정"]
        for c_idx, h in enumerate(headers, start=2):
            cell = ws_sum.cell(row=4, column=c_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border

        for r_idx, m_dict in enumerate(members_data, start=5):
            m = m_dict.get("member", {})
            sec = m_dict.get("section", {})
            dcr = float(m_dict.get("summary_dcr", 0.0))
            is_ok = dcr <= 1.0

            ws_sum.cell(row=r_idx, column=2, value=r_idx - 4).border = self.thin_border
            ws_sum.cell(row=r_idx, column=3, value=m.get("name", f"M-{r_idx}")).border = self.thin_border
            ws_sum.cell(row=r_idx, column=4, value=m.get("type", "RC Member")).border = self.thin_border
            ws_sum.cell(row=r_idx, column=5, value=f"{sec.get('b', 400)}x{sec.get('h', 600)}").border = self.thin_border
            
            c_dcr = ws_sum.cell(row=r_idx, column=6, value=f"{dcr:.3f}")
            c_dcr.border = self.thin_border
            c_dcr.alignment = self.align_right
            c_dcr.font = self.font_pass if is_ok else self.font_fail

            c_stat = ws_sum.cell(row=r_idx, column=7, value="PASS" if is_ok else "NG")
            c_stat.border = self.thin_border
            c_stat.alignment = self.align_center
            c_stat.fill = self.pass_fill if is_ok else self.fail_fill
            c_stat.font = self.font_pass if is_ok else self.font_fail

        # Sheets 2~N: Individual Member Sheets
        for m_dict in members_data:
            m_name = m_dict.get("member", {}).get("name", "Memb")[:28]
            ws_m = wb.create_sheet(title=f"M_{m_name}")
            self._build_checks_sheet(ws_m, m_dict.get("checks", []))

        # Auto-fit columns
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

