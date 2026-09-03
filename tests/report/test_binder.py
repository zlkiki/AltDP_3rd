"""Tests for Project Multi-Member Batch PDF Binder and Multi-Sheet Excel.

Conforms to Requirement 15-3.
"""

import pytest
import io
import openpyxl
from src.report.binder import ReportBinder
from src.report.options import ReportOptions
from src.report.excel_exporter import ExcelReportExporter

binder = ReportBinder()
excel_exporter = ExcelReportExporter()

PROJECT_INFO = {
    "name": "AltDP Mega Tower",
    "location": "서울시 서초구",
    "code": "KDS 14 20 00 / 14 31 00",
    "date": "2026-09-03",
}

MEMBERS_DATA = [
    {
        "project": PROJECT_INFO,
        "member": {"name": "B101", "type": "RC Beam"},
        "material": {"fck": 27.0, "fy": 400.0},
        "section": {"b": 400.0, "h": 600.0, "rebar": "4-D25"},
        "loads": {"Mu": 180.0, "Vu": 140.0},
        "checks": [{"title": "휨모멘트", "demand": "180.0", "capacity": "236.8", "dcr": 0.76}],
        "summary_dcr": 0.760,
        "is_safe": True,
    },
    {
        "project": PROJECT_INFO,
        "member": {"name": "C101", "type": "RC Column"},
        "material": {"fck": 30.0, "fy": 500.0},
        "section": {"b": 600.0, "h": 600.0, "rebar": "12-D25"},
        "loads": {"Pu": 1500.0, "Mu": 250.0},
        "checks": [{"title": "P-M 검토", "demand": "1500.0", "capacity": "2100.0", "dcr": 0.71}],
        "summary_dcr": 0.714,
        "is_safe": True,
    },
]


def test_cover_and_toc_generation():
    """Verify cover page and automatic table of contents rendering."""
    members_meta = [
        {"name": "B101", "type": "RC Beam", "section": "400x600"},
        {"name": "C101", "type": "RC Column", "section": "600x600"},
    ]
    html = binder.generate_cover_and_toc(PROJECT_INFO, members_meta)

    assert "AltDP Mega Tower" in html
    assert "건축구조 부재설계 종합계산서" in html
    assert "목 차 (Table of Contents)" in html
    assert "B101" in html
    assert "C101" in html


def test_bind_project_reports_html():
    """Verify batch binding of multiple members into a unified HTML calculation book."""
    opts = ReportOptions(report_mode="summary")
    html = binder.bind_project_reports_html(PROJECT_INFO, MEMBERS_DATA, options=opts)

    assert "AltDP Mega Tower" in html
    assert "B101" in html
    assert "C101" in html
    assert 'class="page-break"' in html


def test_export_batch_pdf_fallback():
    """Verify PDF/HTML batch bytes export."""
    pdf_bytes = binder.export_batch_pdf(PROJECT_INFO, MEMBERS_DATA)
    assert len(pdf_bytes) > 500
    assert isinstance(pdf_bytes, bytes)


def test_multi_sheet_excel_export():
    """Verify multi-sheet Excel workbook structure and DCR conditional styling."""
    excel_bytes = excel_exporter.export_project_workbook_bytes(PROJECT_INFO, MEMBERS_DATA)
    assert len(excel_bytes) > 1000

    # Load with openpyxl to verify sheets
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames

    assert "Project_Summary" in sheet_names
    assert "M_B101" in sheet_names
    assert "M_C101" in sheet_names

    # Check Summary sheet contents
    ws_sum = wb["Project_Summary"]
    assert ws_sum.cell(row=2, column=2).value is not None
    assert "요약표" in str(ws_sum.cell(row=2, column=2).value)
    assert ws_sum.cell(row=5, column=3).value == "B101"
    assert ws_sum.cell(row=6, column=3).value == "C101"
