"""Unit tests for KDS Quantity Engine and Multi-Sheet Excel Exporter (Phase 17-2)."""

import openpyxl
import pytest

from src.engine.project.quantity_engine import (
    QuantityEngine,
    MemberQuantityInput,
    ProjectQuantitySummary,
    REBAR_SPECS,
)
from src.report.excel_quantity_exporter import ExcelQuantityExporter


def test_development_and_splice_length():
    """Verify KDS 14 20 52 tension development and lap splice formulas."""
    # D22 (db = 22.2 mm), fck = 24 MPa, fy = 400 MPa, bottom bar (alpha = 1.0)
    # ld = (400 / (1.4 * 1.0 * sqrt(24))) * 1.0 * 22.2 = 58.39 * 22.2 = 1296.2 mm
    ld_bot = QuantityEngine.calculate_tension_development_length("D22", fck=24.0, fy=400.0, is_top_bar=False)
    assert 1250.0 < ld_bot < 1350.0

    # Top bar: alpha = 1.3
    ld_top = QuantityEngine.calculate_tension_development_length("D22", fck=24.0, fy=400.0, is_top_bar=True)
    assert abs(ld_top - ld_bot * 1.3) < 1.0

    # Class B splice: 1.3 * ld
    splice = QuantityEngine.calculate_lap_splice_length("D22", fck=24.0, fy=400.0, is_top_bar=False)
    assert abs(splice - 1.3 * ld_bot) < 1.0


def test_member_and_project_quantity_takeoff():
    """Verify concrete, formwork, and rebar tonnage calculations for beams and columns."""
    members = [
        # Beam: b=400, h=600, L=6000
        MemberQuantityInput(
            member_id=1,
            story="2F",
            member_type="BEAM",
            b=400.0,
            h=600.0,
            length=6000.0,
            main_bar_size="D22",
            main_bar_count=8,
            sub_bar_size="D10",
            sub_bar_spacing=200.0
        ),
        # Column: b=600, h=600, L=3500
        MemberQuantityInput(
            member_id=2,
            story="1F",
            member_type="COLUMN",
            b=600.0,
            h=600.0,
            length=3500.0,
            main_bar_size="D25",
            main_bar_count=12,
            sub_bar_size="D10",
            sub_bar_spacing=300.0
        )
    ]

    summary = QuantityEngine.aggregate_project_quantities(members)

    # Beam concrete: 0.4 * 0.6 * 6.0 = 1.44 m3
    # Column concrete: 0.6 * 0.6 * 3.5 = 1.26 m3
    # Total concrete: 2.70 m3
    assert abs(summary.total_concrete_vol_m3 - 2.70) < 0.01

    # Beam formwork: (2*0.6 + 0.4) * 6.0 = 9.6 m2
    # Column formwork: 2*(0.6 + 0.6) * 3.5 = 8.4 m2
    # Total formwork: 18.0 m2
    assert abs(summary.total_formwork_area_m2 - 18.0) < 0.01

    # Rebar tonnage should be positive and reasonable
    assert summary.total_rebar_weight_ton > 0.0
    assert "D22" in summary.rebar_totals_by_size_kg
    assert "D25" in summary.rebar_totals_by_size_kg
    assert "D10" in summary.rebar_totals_by_size_kg

    # Check story breakdown
    assert "1F" in summary.story_breakdowns
    assert "2F" in summary.story_breakdowns
    assert summary.story_breakdowns["2F"]["concrete_m3"] == 1.44
    assert summary.story_breakdowns["1F"]["concrete_m3"] == 1.26


def test_excel_multi_sheet_exporter():
    """Verify generated Excel workbook sheets and structure."""
    members = [
        MemberQuantityInput(member_id=1, story="2F", member_type="BEAM", b=400.0, h=600.0, length=6000.0),
        MemberQuantityInput(member_id=2, story="1F", member_type="COLUMN", b=600.0, h=600.0, length=3500.0),
    ]
    summary = QuantityEngine.aggregate_project_quantities(members)
    stream = ExcelQuantityExporter.export_to_bytes(summary)

    wb = openpyxl.load_workbook(stream)
    sheet_names = wb.sheetnames

    assert "총괄 물량 집계표" in sheet_names
    assert "층별 물량 집계표" in sheet_names
    assert "부재별 상세 내역서" in sheet_names

    ws1 = wb["총괄 물량 집계표"]
    assert "KDS 표준 총괄 공사 물량 집계표 (Summary)" in str(ws1.cell(row=2, column=2).value)
